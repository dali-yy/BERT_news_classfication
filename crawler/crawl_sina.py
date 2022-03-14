# -*- coding: utf-8 -*-
# @Time : 2021/7/8 21:05
# @Author : XXX
# @Site : 
# @File : crawl_sina.py
# @Software: PyCharm
import demjson
import loguru
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm


def get_html_text(url, headers):
    """
    获取网页的html源码
    :param url: 网页地址
    :param head: 请求头
    :return:
    """
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        response.encoding = 'utf-8'
        return response.text
    except:
        return ""


def get_hrefs(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    divs = soup.find_all('div', {"class": "conwp"})
    for div in divs:
        for h3 in div.find_all("h3"):
            href = h3.find('a').get("href")  # 获取链接
            hrefs.append(href)
    return hrefs


def get_hrefs_(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    ul = soup.find('ul', {"class": "new-list"})
    for h4 in ul.find_all('h4'):
        href = h4.find('a').get("href")  # 获取链接
        hrefs.append(href)
    return hrefs


def get_hrefs_2(url, headers):
    hrefs = []
    response = requests.get(url, headers=headers)
    response.encoding = "gbk"
    for item in demjson.decode(response.text[14:-1]):
        hrefs.append(item.get('docurl'))
    return hrefs


def get_title_and_content(url, headers):
    """
    获取新闻的标题和内容
    :param url:
    :param headers:
    :return:
    """
    content = ""  # 新闻内容
    title = ""
    html_text = get_html_text(url, headers)  # 获取网页源码
    soup = BeautifulSoup(html_text, "html.parser")

    title_div = soup.find('div', attrs={"class": 'crticalcontent'})
    if title_div is not None:
        title = title_div.find('h1').string  # 获取新闻标题
    # title_h1 = soup.find("h1", attrs={"class": "l_tit"})
    # if title_h1 is not None:
    #     title = title_h1.text

    content_div = soup.find('div', attrs={"id": "artibody"})
    if content_div is not None:
        for p in content_div.find_all('p'):
            if p.text is not None:
                content += p.text.strip("\n").strip()  # 拼接新闻内容

    return title, content


def crawl_news(url, headers, type, save_path):
    """
    爬取某一类型的新闻
    :param url:
    :param headers:
    :param type:
    :return:
    """
    # 打开存放爬取结果的文件
    wb = load_workbook(save_path)
    ws = wb[type]

    # 添加头部信息
    ws['A1'] = "content"
    ws['B1'] = "channelName"
    ws['C1'] = "title"

    hrefs = get_hrefs(url, headers)  # 获取所有新闻页的网页链接
    for href in tqdm(hrefs[:500]):
        title, content = get_title_and_content(href, headers)
        if title and content:
            ws.append([content, type, title])

    wb.save(save_path)  # 保存
    loguru.logger.success("url:" + url + " 爬取成功！")


if __name__ == "__main__":
    headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                             "Chrome/91.0.4472.114 Safari/537.36"}
    url = "https://interface.sina.cn/pc_api/public_news_data.d.json?callback=jQuery11120807885976026907_1625880456877" \
          "&cids=209211&pdps=PDPS000000060130%2CPDPS000000066866&smartFlow=&type=std_news%2Cstd_slide%2Cstd_video" \
          "&pageSize=20&top_id=hencxtu1691422%2Chencxtu5974075%2Chencxtu5919005%2Chencxtu5908111&mod=nt_culture0" \
          "&cTime=1483200000&up=0&action=0&_=1625880456878 "
    # for i in range(1, 10):
    #     if i == 1:
    #         url = "https://tech.163.com/internet/"
    #     else:
    #         url = f"https://tech.163.com/special/internet_2016_0{i}/"  # 央广网
    #     crawl_news(url, headers, "科技", "./news/新浪新闻.xlsx")
    # print(len(get_hrefs(url, headers)))
    # print(len(get_hrefs_2(url, headers)))
    # title, content = get_title_and_content(url, headers)
    # print(title)
    # print(content)
    # crawl_news(url, headers, "游戏", "../dataset/新浪新闻.xlsx")
    response = requests.get(url, headers=headers)
    # response.encoding = "utf-8"
    # print(response.text)
    for item in demjson.decode(response.text[14:-1]):
        print(item.get('url'))