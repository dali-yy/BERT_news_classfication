# -*- coding: utf-8 -*-
# @Time : 2021/7/7 19:49
# @Author : XXX
# @Site : 
# @File : crawl_tencent.py
# @Software: PyCharm
# 爬取人民网新闻
import re

import loguru
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup


def get_html_text(url, headers):
    """
    获取网页的html源码
    :param url: 网页地址
    :param head: 请求头
    :return:
    """
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        response.encoding = 'gbk'
        return response.text
    except:
        return ""


def get_hrefs(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    uls = soup.find_all('ul', {"class": "list_16 mt10"})
    for ul in uls:
        _as = ul.find_all('a')
        for a in _as:
            href = a.get("href")  # 获取链接
            if not href.startswith("http"):
                href = re.match(r'http.*?cn', url).group() + href
            hrefs.append(href)
    return hrefs


def get_title_and_content(url, headers):
    """
    获取新闻的标题和内容
    :param url:
    :param headers:
    :return:
    """
    content = ""  # 新闻内容
    title = ""
    html_text = get_html_text(url, headers)  # 获取网页源码
    soup = BeautifulSoup(html_text, "html.parser")

    title_div = soup.find('div', attrs={"class": "col col-1 fl"})
    if title_div is not None:
        title = title_div.find('h1').string  # 获取新闻标题

    content_div = soup.find('div', attrs={"class": "rm_txt_con cf"})
    if content_div is not None:
        for p in content_div.find_all('p'):
            if p.string is not None:
                content += p.string.strip("\n").strip()  # 拼接新闻内容

    return title, content


def crawl_news(url, headers, type, save_path):
    """
    爬取某一类型的新闻
    :param url:
    :param headers:
    :param type:
    :return:
    """
    # 打开存放爬取结果的文件
    wb = load_workbook(save_path)
    ws = wb[type]

    # 添加头部信息
    ws['A1'] = "content"
    ws['B1'] = "channelName"
    ws['C1'] = "title"

    hrefs = get_hrefs(url, headers)  # 获取所有新闻页的网页链接
    for href in hrefs:
        title, content = get_title_and_content(href, headers)
        if title and content:
            ws.append([content, type, title])

    wb.save(save_path)  # 保存
    loguru.logger.success("url:" + url + " 爬取成功！")


if __name__ == "__main__":

    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                             "Chrome/91.0.4472.114 Safari/537.36"}
    for i in range(1, 11):
        url = f"http://legal.people.com.cn/GB/205462/index{i}.html"  # 人民网新闻网站地址
        crawl_news(url, headers, "其他", "./news/新闻.xlsx")
    # url = "http://travel.people.com.cn/GB/41636/index.html"
    # crawl_news(url, headers, "娱乐", "./news/人民网.xlsx")

