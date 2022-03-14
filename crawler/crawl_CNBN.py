# -*- coding: utf-8 -*-
# @Time : 2021/7/8 14:49
# @Author : XXX
# @Site : 
# @File : crawl_CNBN.py
# @Software: PyCharm
# 分类爬取央广网新闻

import loguru
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup


def get_html_text(url, headers):
    """
    获取网页的html源码
    :param url: 网页地址
    :param head: 请求头
    :return:
    """
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        response.encoding = 'gbk'
        return response.text
    except:
        return ""


def get_hrefs(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    divs = soup.find_all('div', {"class": "item"})
    for div in divs:
        a = div.find('a')
        href = a.get("href")  # 获取链接
        hrefs.append(href)
    return hrefs


def get_hrefs_(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    uls = soup.find_all('ul', {"class": "topnews_nlist"})
    for ul in uls:
        for h3 in ul.find_all('h3'):
            href = h3.find('a').get("href")  # 获取链接
            hrefs.append(href)
    return hrefs


def get_title_and_content(url, headers):
    """
    获取新闻的标题和内容
    :param url:
    :param headers:
    :return:
    """
    content = ""  # 新闻内容
    title = ""
    html_text = get_html_text(url, headers)  # 获取网页源码
    soup = BeautifulSoup(html_text, "html.parser")

    title_div = soup.find('div', attrs={"class": "article-header"})
    if title_div is not None:
        title = title_div.find('h1').string  # 获取新闻标题
    # title_h1 = soup.find("h1", attrs={"class": "f24 lh40 fb txtcenter f12_292929 yahei"})
    # if title_h1 is not None:
    #     title = title_h1.text

    content_div = soup.find('div', attrs={"class": "article-body"})
    if content_div is not None:
        for p in content_div.find_all('p'):
            if p.string is not None:
                content += p.string.strip("\n").strip()  # 拼接新闻内容

    return title, content


def crawl_news(url, headers, type, save_path):
    """
    爬取某一类型的新闻
    :param url:
    :param headers:
    :param type:
    :return:
    """
    # 打开存放爬取结果的文件
    wb = load_workbook(save_path)
    ws = wb[type]

    # 添加头部信息
    ws['A1'] = "content"
    ws['B1'] = "channelName"
    ws['C1'] = "title"

    hrefs = get_hrefs(url, headers)  # 获取所有新闻页的网页链接
    for href in hrefs:
        title, content = get_title_and_content(href, headers)
        if title and content:
            ws.append([content, type, title])

    wb.save(save_path)  # 保存
    loguru.logger.success("url:" + url + " 爬取成功！")


if __name__ == "__main__":
    headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                             "Chrome/91.0.4472.114 Safari/537.36"}
    url = "https://sports.163.com/"
    # for i in range(0, 30):
    #     if i == 0:
    #         url = "http://ent.cnr.cn/zx/index.html"
    #     else:
    #         url = f"http://ent.cnr.cn/zx/index_{i}.html"  # 央广网
    #     crawl_news(url, headers, "其他", "./news/央广网.xlsx")
    print(get_hrefs_(url, headers))
    # title, content = get_title_and_content(url, headers)
    # print(title)
    # print(content)
    # crawl_news(url, headers, "教育", "./news/央广网.xlsx")
