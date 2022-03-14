# -*- coding: utf-8 -*-
# @Time : 2021/7/8 16:55
# @Author : XXX
# @Site : 
# @File : craw_wangyi.py
# @Software: PyCharm
import demjson
import loguru
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm


def get_html_text(url, headers):
    """
    获取网页的html源码
    :param url: 网页地址
    :param headers: 请求头
    :return:
    """
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        # response.encoding = 'utf-8'
        return response.text
    except:
        return ""


def get_hrefs(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    divs = soup.find_all('div', {"class": "news_main_info"})
    for div in divs:
        href = div.find('h2').find('a').get("href")  # 获取链接
        hrefs.append(href)
    return hrefs


def get_hrefs_(url, headers):
    """
    获取所有新闻链接
    :param url:
    :param headers:
    :return:
    """
    hrefs = []
    html_text = get_html_text(url, headers)  # 获取新闻列表页网页源码
    soup = BeautifulSoup(html_text, "html.parser")  # bs4解析源码
    ul = soup.find('ul', {"class": "house_text_list"})
    for h3 in ul.find_all('h3'):
        href = h3.find('a').get("href")  # 获取链接
        hrefs.append(href)
    return hrefs


def get_hrefs_2(url, headers):
    reponse = requests.get(url, headers=headers)
    reponse.encoding = "gbk"
    result = reponse.text.split('{')
    hrefs = []
    for r in result[10:]:
        hrefs.append(r.split(',')[2][3:-1])
    return hrefs


def get_hrefs_3(url, headers):
    hrefs = []
    response = requests.get(url, headers=headers)
    response.encoding = "gbk"
    for item in demjson.decode(response.text[14:-1]):
        hrefs.append(item.get('docurl'))
    return hrefs


def get_title_and_content(url, headers):
    """
    获取新闻的标题和内容
    :param url:
    :param headers:
    :return:
    """
    content = ""  # 新闻内容
    title = ""
    html_text = get_html_text(url, headers)  # 获取网页源码
    soup = BeautifulSoup(html_text, "html.parser")

    title_h1 = soup.find("h1", attrs={"class": "post_title"})
    if title_h1 is not None:
        title = title_h1.text

    content_div = soup.find('div', attrs={"class": "post_body"})
    if content_div is not None:
        for p in content_div.find_all('p'):
            if p.text is not None:
                content += p.text.strip("\n").strip()  # 拼接新闻内容

    return title, content


def crawl_news(url, headers, type, save_path):
    """
    爬取某一类型的新闻
    :param url:
    :param headers:
    :param type:
    :return:
    """
    # 打开存放爬取结果的文件
    wb = load_workbook(save_path)
    ws = wb[type]

    # 添加头部信息
    ws['A1'] = "content"
    ws['B1'] = "channelName"
    ws['C1'] = "title"

    hrefs = get_hrefs_(url, headers)  # 获取所有新闻页的网页链接
    for href in tqdm(hrefs):
        title, content = get_title_and_content(href, headers)
        if title and content:
            ws.append([content, type, title])

    wb.save(save_path)  # 保存
    loguru.logger.success("url:" + url + " 爬取成功！")


if __name__ == "__main__":
    headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) "
                             "Chrome/91.0.4472.114 Safari/537.36"}
    # url = "http://nj.house.163.com/special/020305CI/fqdb1.html"
    for i in range(1, 21):
        if i == 1:
            url = "http://nj.house.163.com/special/020305CI/lszn1.html"
        elif i < 10:
            url = f"http://nj.house.163.com/special/020305CI/lszn1_0{i}.html"
        else:
            url = f"http://nj.house.163.com/special/020305CI/lszn1_{i}.html"
        crawl_news(url, headers, "房产", "./news/新闻.xlsx")
    # print(len(get_hrefs(url, headers)))
    # print(len(get_hrefs_(url, headers)))
    # print(len(get_hrefs_3(url, headers)))
    # title, content = get_title_and_content(url, headers)
    # print(title)
    # print(content)
    # crawl_news(url, headers, "其他", "./news/新闻.xlsx")
